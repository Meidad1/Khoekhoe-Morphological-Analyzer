import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class GlossingDictGenerator:
    MORPHEME_IDX = 1
    GE_IDX = 2
    PS_IDX = 3

    def __generate_dict(self, table: Worksheet):
        """
        :return: A dict of the form: {tx : [ge, ps]}
        """
        cur_dict = {}
        for i in range(2, table.max_row):
            cur_dict[table.cell(i, self.MORPHEME_IDX).value] = \
                [table.cell(i, self.GE_IDX).value, table.cell(i, self.PS_IDX).value]     # morpheme/lexeme : [gloss, pos]
        return cur_dict

    def get_dict(self, excel_path: str):
        book = openpyxl.load_workbook(excel_path)
        table = book.active
        return self.__generate_dict(table)
