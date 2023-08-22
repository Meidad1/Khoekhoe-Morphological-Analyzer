import re
from pprint import pprint

from lemma import Lemma
from openpyxl.worksheet.worksheet import Worksheet


class glossesDictionaryParser:

    def __init__(self):
        self.adverbs = set()

    def parse_dicts(self, table: Worksheet):
        """
        :return: Two dicts (grammatical and lexical) of the form: {mb : [lemma1, lemma2, ...]}
        Each key is linked to a list of possible pairs (tuples) for glossing.
        """
        headers_to_idx_map = self.map_headers_to_indexes(table)
        gram_dict = {}
        lex_dict = {}
        for i in range(2, table.max_row + 1):
            if table.cell(i, headers_to_idx_map['Ignore?']).value != "yes":
                mb = table.cell(i, headers_to_idx_map['mb']).value
                ge = str(table.cell(i, headers_to_idx_map['ge']).value)
                ps = str(table.cell(i, headers_to_idx_map['ps']).value)
                gender = table.cell(i, headers_to_idx_map['Gender']).value
                so = table.cell(i, headers_to_idx_map['Source']).value
                other_trans = table.cell(i, headers_to_idx_map['Other Translations']).value
                other_trans_lst = self.parse_other_trans_field(other_trans)
                common_misspellings = str(table.cell(i, headers_to_idx_map['Common Misspellings']).value)
                lemma_type = table.cell(i, headers_to_idx_map['Gram/Lex']).value
                if not mb and ge == "None":  # if it's an empty line -> ignore it
                    continue
                elif not mb or ge == "None" or ps == "None" or not lemma_type or lemma_type not in {"L", "G"} \
                        or (gender is not None and gender not in {"m", "f", "c"}) \
                        or (so is not None and so not in {"Afrikaans", "English"}):
                    raise Exception("Line " + str(i) + " in the glossing dictionary is invalid.")

                if ps == "adv":
                    self.adverbs.add(mb)

                cur_lemma = Lemma(mb, ge, ps, common_misspellings,
                                  gender, so, other_trans_lst, lemma_type)
                if lemma_type == "G":  # if it's a grammatical item
                    if mb not in gram_dict:
                        gram_dict[mb] = [cur_lemma]  # create a new list for cur_mb and add its lemma object
                    else:
                        gram_dict[mb].append(cur_lemma)  # add the new possible sense to the existing list of cur_mb
                else:  # if it's a lexical item
                    if mb not in lex_dict:
                        lex_dict[mb] = [cur_lemma]
                    else:
                        lex_dict[mb].append(cur_lemma)

        return gram_dict, lex_dict

    def map_headers_to_indexes(self, table: Worksheet) -> dict:
        headers_to_idx_map = dict()
        for j in range(1, table.max_column + 1):
            cur_header = table.cell(1, j).value
            if cur_header:
                headers_to_idx_map[cur_header] = j
        return headers_to_idx_map

    def get_adverbs_set(self):
        return self.adverbs

    def parse_other_trans_field(self, other_trans: str) -> list:
        res = []
        if other_trans:
            parentheses_pattern = r"\([^()]*\)"
            other_trans = re.sub(parentheses_pattern, "", other_trans)
            other_trans = re.sub("  ", " ", other_trans).strip()
            res = re.split(r'[,;]', other_trans)
            res = [item.strip() for item in res]
        return res
